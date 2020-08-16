#!/usr/bin/python
# -*- coding: UTF-8 -*-
# @Created on :  2020/8/9
# @Created by :  Stevie

from f5.bigip import ManagementRoot
from openpyxl import Workbook
import os
import re

def fun_ping_mod(get_host_ip):
    result_ping = os.system('ping -c 1 -W 2 %s' % get_host_ip)
    # print(result_ping)
    return result_ping

pool_node_link_list_0 = []
pool_node_link_list = []
vs_pool_link_list_0 = []
vs_pool_link_list = []
vs_pool_node_link_list_0 = []
vs_pool_node_link_list = []
rule_list = []
rule_info_list = []
vs_rule_link_list = []
vs_rule_node_link_list = []

wb = Workbook()
dest_file = 'f5.xlsx'

f5_login_info = ['1.1.1.1', 'web_admin', 'web_YBwonders2008', '10']
f5_mgmt = ''

# 连接
ping_test = fun_ping_mod(f5_login_info[0])
if ping_test == 0:
    print('F5 dev is available.')
    try:
        f5_mgmt = ManagementRoot(f5_login_info[0], f5_login_info[1], f5_login_info[2], timeout=f5_login_info[3])
        print('login success')
    except BaseException as error_0:
        print(error_0)
        print('er0')
else:
    print('Ping 测试失败. 请检查IP: ' + f5_login_info[0] + ' 状态...')

pools = f5_mgmt.tm.ltm.pools.get_collection()
for pool in pools:
    # print(pool.name)
    for node_in_pool in pool.members_s.get_collection():
        # print(node_in_pool.name)
        pool_node_link_list_0 = [pool.name, node_in_pool.name]
        pool_node_link_list.append(pool_node_link_list_0)
pool_node_link_list_0 = []
# print(pool_node_link_list)

virtual_servers = f5_mgmt.tm.ltm.virtuals.get_collection()
for virtual_server in virtual_servers:
    # print(virtual_server.raw)
    # print(virtual_server.destination)
    try:
        vs_pool_link_list_0 = [virtual_server.name, virtual_server.destination.split('/')[-1],
                               virtual_server.pool.split('/')[-1]]
        vs_pool_link_list.append(vs_pool_link_list_0)
        # print(virtual_server.pool)
    except AttributeError as error_1:
        try:
            # pass
            # print(virtual_server.name)
            # print(virtual_server.destination.split('/')[-1])
            tmp_rule = str(virtual_server.rules)
            # tmp_rule = tmp_rule.split('/')[-1].rstrip("']")
            # print(tmp_rule)
            vs_rule_link_list_0 = [virtual_server.name,  virtual_server.destination.split('/')[-1],
                                   tmp_rule.split('/')[-1].rstrip("']")]
            vs_rule_link_list.append(vs_rule_link_list_0)
            # print(virtual_server.rules)
            # print(vs_rule_link_list_0)
        except AttributeError as error_2:
            print('出错了?')

for x in vs_pool_link_list:
    for y in pool_node_link_list:
        if x[2] == y[0]:
            vs_pool_node_link_list_0 = (x[0], x[1], y[0], y[1])
            vs_pool_node_link_list.append(vs_pool_node_link_list_0)
        else:
            pass




rule_list = f5_mgmt.tm.ltm.rules.get_collection()
for rule_single in rule_list:
    # print(i.name)
    if 'iRule' in rule_single.name:
        rule = rule_single.apiAnonymous.replace('\n', '')
        rule = rule.replace('\t', '')
        rule = rule.replace(' ', '')
        rule = rule.replace('whenHTTP_REQUEST{if{', '')
        rule = rule.replace('}{pool', '}{pool:')
        # print(rule)
        rule_1 = rule.split('elseif{')
        rule_else = rule.replace('elseif', '')
        rule_else = rule_else.split('else{')
        # print(rule_else)
        if 'HTTP' not in rule_else[-1]:
            rule_else_info = rule_else[-1]
            print(rule_else_info)
        else:
            print('2')
        print('1')
        # 提取url
        for rule_info_0 in rule_1:
            r = re.search(r"(\"[a-z]*\"|\"[/,a-z,A-Z,0-9,_]*\")", rule_info_0)
            if r:
                r_0 = r.group().replace('/', '')
                r_0 = r_0.rstrip('"')
                r_0 = r_0.replace('"', '/')
            else:
                pass
            rr = re.search(r"(pool:[a-z,A-Z,0-9,_]*)", rule_info_0)
            if rr:
                r_1 = rr.group()
                r_1 = r_1.lstrip('pool:')
            else:
                pass
            single = [rule_single.name, r_0, r_1]
            rule_info_list.append(single)
# print(rule_info_list)

# [virtual_server.name,  virtual_server.destination.split('/')[-1],
#                                    tmp_rule.split('/')[-1]]

for x in vs_rule_link_list:
    # print(x)
    for y in rule_info_list:
        # print(y)
        if x[2] == y[0]:
            for z in pool_node_link_list:
                # print(z)
                if y[2] == z[0]:
                    vs_rule_node_link_list_0 = (x[0], x[1], y[0], y[1], y[2], z[1])
                    vs_rule_node_link_list.append(vs_rule_node_link_list_0)
                    print(vs_rule_node_link_list_0)
        else:
            pass
# print(vs_rule_node_link_list)


# f5_info_file = open('f5_info_file.txt', 'w', encoding='utf-8-sig')
# for info in vs_pool_node_link_list:
#     f5_info_file.write('VS名称: ' + info[0] + ' VS_IP地址: ' + info[1] +
#                        ' 关联Pool名称: ' + info[2] + ' 后端IP地址: ' + info[3] + '\n')
#     print(info)
# f5_info_file.close()
# print(':) Py by Stevie_Chen...')

list_len = len(vs_pool_node_link_list)
ws1 = wb.active
ws1.title = 'F5_VS_Link_Info'
xls_head = ['VS名称', 'VS_IP地址', '关联Pool名称', '后端IP地址', '关联url', '关联irule名称']
ws_row = 2

for col in range(1, 7):
    ws1.cell(column=col, row=1, value=xls_head[col-1])

for info in vs_pool_node_link_list:
    ws1.cell(column=1, row=ws_row, value=info[0])
    ws1.cell(column=2, row=ws_row, value=info[1])
    ws1.cell(column=3, row=ws_row, value=info[2])
    ws1.cell(column=4, row=ws_row, value=info[3])
    ws_row = ws_row + 1
    # print(info)
for info in vs_rule_node_link_list:
    ws1.cell(column=1, row=ws_row, value=info[0])
    ws1.cell(column=2, row=ws_row, value=info[1])
    ws1.cell(column=3, row=ws_row, value=info[4])
    ws1.cell(column=4, row=ws_row, value=info[5])
    ws1.cell(column=5, row=ws_row, value=info[3])
    ws1.cell(column=6, row=ws_row, value=info[2])
    ws_row = ws_row + 1

ws1.column_dimensions['A'].width = 30.0
ws1.column_dimensions['B'].width = 23.0
ws1.column_dimensions['C'].width = 32.0
ws1.column_dimensions['D'].width = 23.0
ws1.column_dimensions['E'].width = 19.0
ws1.column_dimensions['F'].width = 23.0

wb.save(filename=dest_file)

print(':) Py by Stevie_Chen...')

# 待更新 'i_rule_else'部分
#
# print(vs_pool_node_link_list)
#
# # pool name
# for i in pool_node_link_list:
#     print(i[0])
# # node name
# for i in pool_node_link_list:
#     print(i[1])