#!/usr/bin/python
# -*- coding: UTF-8 -*-
# @Time       :  2020/8/4 14:04
# @Created by :  Stevie

from netmiko import ConnectHandler, NetmikoTimeoutException, NetmikoAuthenticationException  # 引入连接模块、报错模块
import time
from openpyxl import load_workbook
import threading
import os


def directory_check_mod(local_path_mod, directory_name_mod):
    path_mod = os.listdir(local_path_mod)
    if directory_name_mod not in path_mod:
        os.mkdir(local_path_mod + '/' + directory_name_mod)
    else:
        print(local_path_mod + '/' + directory_name_mod + '已存在.')

def file_mod(get_file_name, get_ip_info, get_info, get_path):
    # file_creat = open(config_path +
    # str(get_file_name) + '_config_backup_at_' + time_str + '.txt', 'w', encoding='utf-8-sig')
    file_creat = open(get_path + str(get_file_name) + '.txt', 'w', encoding='utf-8-sig')
    file_creat.write('dev_ip: ' + get_ip_info + ' created_at: ' + time_str + '\n')
    file_creat.write('*' * 85 + '\n')
    file_creat.write(get_info + '\n')
    file_creat.close()

def fun_ping_mod(get_host_ip):
    result_ping = os.system('ping -c 1 -W 2 %s' % get_host_ip)
    # print(result_ping)
    return result_ping

def my_ssh_mod(dev_type_mod, dev_host_ip_mod, dev_username_mod, dev_password_mod, dev_enable_pass_mod, dev_module_mod):
    with th_limit:
        ping_test = fun_ping_mod(dev_host_ip_mod)
        if ping_test == 0:
            net_dev_host = {
                'device_type': dev_type_mod,
                'host': dev_host_ip_mod,
                'username': dev_username_mod,
                'password': dev_password_mod,
                'secret': dev_enable_pass_mod,
                'timeout': 5
            }
            try:
                net_connect = ConnectHandler(**net_dev_host)
                net_connect.enable()
                # 获取设备模块信息
                if dev_module_mod == 'module':
                    hw_state = net_connect.send_command('show module')
                    print(str(net_dev_host['host']) + ' --- 模块状态获取完成.')
                    file_mod(net_dev_host['host'], net_dev_host['host'], hw_state, module_path)
                    # dev_info_result.append([net_dev_host['host'], hw_state])
                elif dev_module_mod == 'platform':
                    hw_state = net_connect.send_command('show platform')
                    print(str(net_dev_host['host']) + ' --- 模块状态获取完成.')
                    file_mod(net_dev_host['host'], net_dev_host['host'], hw_state, module_path)
                    # dev_info_result.append([net_dev_host['host'], hw_state])
                else:
                    print(str(net_dev_host['host']) + ' --- 非模块化设备无需检查.')
                print(str(net_dev_host['host']) + ' --- 模块状态记录完成.')
                # 获取设备show run
                if dev_type_mod == 'huawei':
                    dev_config = net_connect.send_command('dis curr')
                    print(str(net_dev_host['host']) + ' --- 配置获取完成.')
                else:
                    dev_config = net_connect.send_command('show run')
                    print(str(net_dev_host['host']) + ' --- 配置获取完成.')
                net_connect.disconnect()
                print(str(net_dev_host['host']) + ' --- SSH连接断开.')

                file_mod(net_dev_host['host'], net_dev_host['host'], dev_config, config_path)

                # print("当前线程数: " + str(threading.active_count()))

                print(str(net_dev_host['host']) + ' --- 配置文件备份完成.')
            except NetmikoTimeoutException as error_info_1:
                print(str(error_info_1))
            except NetmikoAuthenticationException as error_info_2:
                print(str(error_info_2))
        else:
            print('Ping 测试失败. 请检查IP: ' + dev_host_ip_mod + ' 状态...')
    # return dev_info_result

timestamp_start = int(time.time())

act_time = time.localtime()
time_str = (time.strftime('%m-%d_%H-%M', act_time))
print('当前时间: ' + time_str)

need_path = '.'
directory_name_list = ['for_module', 'for_config', 'for_test']  # [0]存放dev_module [1]存放show run [2]待定

for directory_name in directory_name_list:
    if directory_name:
        directory_check_mod(need_path, directory_name)
    else:
        print('无文件夹名.')

module_path = need_path + '/' + str(directory_name_list[0]) + '/'
config_path = need_path + '/' + str(directory_name_list[1]) + '/'

date = time.strftime('%Y%m%d', time.localtime())  # 赋予date变量
xls_name = load_workbook('dev_list.xlsx')
xls_worksheet = xls_name["Sheet1"]
xls_name.active = xls_worksheet
# net_dev_list_info = []
threads = []
th_limit = threading.Semaphore(10)  # 线程数

# dev_info_result = []

for vaule in xls_worksheet.iter_rows(min_row=2, max_col=6, values_only=True):
    info_from_xls = list(vaule)
    dev_type = str(info_from_xls[0])
    dev_host_ip = str(info_from_xls[1])
    dev_username = str(info_from_xls[2])
    dev_password = str(info_from_xls[3])
    dev_enable_pass = str(info_from_xls[4])
    dev_module_state = str(info_from_xls[5])
    # 多线程执行
    th_queue = threading.Thread(target=my_ssh_mod, args=(dev_type, dev_host_ip, dev_username,
                                                         dev_password, dev_enable_pass, dev_module_state))
    th_queue.start()
    threads.append(th_queue)

xls_name.close()

for i in threads:
    i.join()
    # result.append(th_queue.dev_info_result)

# r2file = open(module_path + 'dev_module.txt', 'w', encoding='utf-8-sig')
# for info in dev_info_result:
#     r2file.write(info[0] + '\n')
#     r2file.write('*' * 85 + '\n')
#     r2file.write(info[1] + '\n')
#     r2file.write('*' * 85 + '\n')
# r2file.close()

timestamp_end = int(time.time())
timeuse = (timestamp_end - timestamp_start)

print('多线程共耗时: ' + str(timeuse) + ' seconds.')
print(':) Py by Stevie_Chen...')
